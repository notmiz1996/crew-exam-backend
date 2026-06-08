# questions/admin.py

"""
题库 Admin 配置
- QuestionType：简单的列表展示
- Chapter：树形展示 + 自动计算层级 + 批量导入
- Question：按题型/章节筛选 + 搜索 + 表单动态选项 + 批量导入
"""
from django.contrib import admin
from django import forms
from django.utils.html import format_html
from django.urls import path, reverse
from django.shortcuts import render
from django.contrib import messages
from django.http import HttpResponseRedirect

from .models import QuestionType, Chapter, Question

# ─── 批量导入依赖 ─────────────────────────────
try:
    import openpyxl
except ImportError:
    openpyxl = None


@admin.register(QuestionType)
class QuestionTypeAdmin(admin.ModelAdmin):
    list_display = ('name', 'code', 'score', 'question_count')
    search_fields = ('name', 'code')
    list_editable = ('score',)

    def question_count(self, obj):
        return obj.question_set.count()
    question_count.short_description = '题目数量'


# ═══════════════════════════════════════════════
# Chapter Admin — 章节管理 + 批量导入章节
# ═══════════════════════════════════════════════

@admin.register(Chapter)
class ChapterAdmin(admin.ModelAdmin):
    change_list_template = 'admin/questions/chapter/change_list.html'

    list_display = ('display_with_indent', 'level_depth', 'sort_order', 'question_count')
    list_editable = ('sort_order',)
    search_fields = ('name',)
    list_filter = ('level_depth',)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-chapters/',
                self.admin_site.admin_view(self.import_chapters_view),
                name='questions_chapter_import',
            ),
        ]
        return custom_urls + urls

    def import_chapters_view(self, request):
        """批量导入章节（已有代码，保持不变）"""
        context = {
            'title': '批量导入章节',
            'opts': self.model._meta,
            'has_change_permission': self.has_change_permission(request),
            'site_header': self.admin_site.site_header,
            'site_title': self.admin_site.site_title,
        }

        if openpyxl is None:
            messages.error(request, '缺少 openpyxl 库。请运行: pip install openpyxl')
            return render(request, 'admin/questions/chapter/import_chapters.html', context)

        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, '请选择一个 xlsx 文件上传')
                return render(request, 'admin/questions/chapter/import_chapters.html', context)

            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, '仅支持 .xlsx 格式文件')
                return render(request, 'admin/questions/chapter/import_chapters.html', context)

            try:
                wb = openpyxl.load_workbook(excel_file, read_only=True)
                ws = wb.active
            except Exception as e:
                messages.error(request, f'文件读取失败: {e}')
                return render(request, 'admin/questions/chapter/import_chapters.html', context)

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                messages.warning(request, '文件中没有数据行（仅有表头）')
                return render(request, 'admin/questions/chapter/import_chapters.html', context)

            header = [str(c or '').strip() for c in rows[0]]
            if '章' not in header:
                messages.error(request, f'表头格式不符合要求，请确保包含列：序号、章、节、小节。当前表头: {", ".join(header)}')
                return render(request, 'admin/questions/chapter/import_chapters.html', context)

            expected_headers = ['序号', '章', '节', '小节']
            col_idx = {col: idx for idx, col in enumerate(expected_headers) if col in header}

            imported_count = 0
            skipped_rows = []
            node_cache = {}

            for row_idx, row in enumerate(rows[1:], start=2):
                if all(c is None or str(c).strip() == '' for c in row):
                    continue

                serial = str(row[col_idx.get('序号', 0)] or '').strip() if '序号' in col_idx else ''
                zhang = str(row[col_idx['章']] or '').strip() if '章' in col_idx else ''
                jie = str(row[col_idx.get('节', col_idx.get('章') + 1)] or '').strip() if '节' in col_idx else ''
                xiaojie = str(row[col_idx.get('小节', col_idx.get('节', col_idx.get('章') + 1) + 1)] or '').strip() if '小节' in col_idx else ''

                if not zhang:
                    skipped_rows.append(f'第{row_num}行：章名称为空，跳过')
                    continue

                try:
                    # 第1级：章
                    zhang_key = (None, zhang)
                    if zhang_key in node_cache:
                        zhang_node = node_cache[zhang_key]
                    else:
                        zhang_node, created = Chapter.objects.get_or_create(
                            parent=None, name=zhang,
                            defaults={'sort_order': row_num, 'level_depth': 1},
                        )
                        node_cache[zhang_key] = zhang_node
                        if created:
                            imported_count += 1
                        if serial:
                            try:
                                zhang_node.sort_order = int(serial)
                                zhang_node.save(update_fields=['sort_order'])
                            except ValueError:
                                pass

                    if not jie:
                        continue

                    # 第2级：节
                    jie_key = (zhang_node.pk, jie)
                    if jie_key in node_cache:
                        jie_node = node_cache[jie_key]
                    else:
                        jie_node, created = Chapter.objects.get_or_create(
                            parent=zhang_node, name=jie,
                            defaults={'sort_order': row_num, 'level_depth': 2},
                        )
                        node_cache[jie_key] = jie_node
                        if created:
                            imported_count += 1

                    if not xiaojie:
                        continue

                    # 第3级：小节
                    xiaojie_key = (jie_node.pk, xiaojie)
                    if xiaojie_key in node_cache:
                        continue

                    xiaojie_node, created = Chapter.objects.get_or_create(
                        parent=jie_node, name=xiaojie,
                        defaults={'sort_order': row_num, 'level_depth': 3},
                    )
                    node_cache[xiaojie_key] = xiaojie_node
                    if created:
                        imported_count += 1

                except Exception as e:
                    skipped_rows.append(f'第{row_num}行（{zhang} > {jie or ""} > {xiaojie or ""}）导入出错: {e}')
                    continue

            wb.close()

            if imported_count > 0:
                messages.success(request, f'✅ 成功导入 {imported_count} 个章节！')
            if skipped_rows:
                messages.warning(request, f'⚠️ 以下 {len(skipped_rows)} 行被跳过：')
                for msg in skipped_rows[:10]:
                    messages.warning(request, msg)
                if len(skipped_rows) > 10:
                    messages.warning(request, f'... 还有 {len(skipped_rows) - 10} 行被跳过，请检查文件内容')

            return HttpResponseRedirect(reverse('admin:questions_chapter_changelist'))

        return render(request, 'admin/questions/chapter/import_chapters.html', context)

    def display_with_indent(self, obj):
        indent = '　' * (obj.level_depth - 1) if obj.level_depth > 1 else ''
        return f'{indent}├ {obj.name}'
    display_with_indent.short_description = '章节名称'
    display_with_indent.admin_order_field = 'level_depth__sort_order'

    def question_count(self, obj):
        return obj.question_set.count()
    question_count.short_description = '题目数量'


# ═══════════════════════════════════════════════
# Question 表单
# ═══════════════════════════════════════════════

class QuestionForm(forms.ModelForm):
    """
    题目表单：根据题型动态显示选项输入区
    """
    option_a = forms.CharField(label='选项 A', required=False, max_length=500)
    option_b = forms.CharField(label='选项 B', required=False, max_length=500)
    option_c = forms.CharField(label='选项 C', required=False, max_length=500)
    option_d = forms.CharField(label='选项 D', required=False, max_length=500)
    option_e = forms.CharField(label='选项 E', required=False, max_length=500,
                               help_text='多选题专用')
    option_f = forms.CharField(label='选项 F', required=False, max_length=500,
                               help_text='多选题专用')

    class Meta:
        model = Question
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance.pk and self.instance.options:
            opts = self.instance.options
            if len(opts) >= 1:
                self.fields['option_a'].initial = opts[0]
            if len(opts) >= 2:
                self.fields['option_b'].initial = opts[1]
            if len(opts) >= 3:
                self.fields['option_c'].initial = opts[2]
            if len(opts) >= 4:
                self.fields['option_d'].initial = opts[3]
            if len(opts) >= 5:
                self.fields['option_e'].initial = opts[4]
            if len(opts) >= 6:
                self.fields['option_f'].initial = opts[5]

    def clean(self):
        cleaned_data = super().clean()
        question_type = cleaned_data.get('question_type')

        options = []
        for letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            value = cleaned_data.get(f'option_{letter.lower()}')
            if value:
                options.append(f'{letter}. {value}')

        if not question_type:
            return cleaned_data

        if question_type.code in ('single_choice',):
            if len(options) < 2:
                raise forms.ValidationError('单选题至少需要填写 A、B 两个选项')
            cleaned_data['options'] = options[:4]

        elif question_type.code in ('multi_choice',):
            if len(options) < 2:
                raise forms.ValidationError('多选题至少需要填写 A、B 两个选项')
            cleaned_data['options'] = options[:6]

        elif question_type.code in ('judgment',):
            cleaned_data['options'] = ['A. 正确', 'B. 错误']

        return cleaned_data


# ═══════════════════════════════════════════════
# Question Admin — 题目管理 + 批量导入题目
# ═══════════════════════════════════════════════

@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    form = QuestionForm
    change_list_template = 'admin/questions/question/change_list.html'

    list_display = ('id', 'stem_short', 'question_type', 'chapter', 'answer_display', 'created_at')
    list_filter = ('question_type', 'chapter',)
    search_fields = ('stem',)
    list_display_links = ('id', 'stem_short')
    date_hierarchy = 'created_at'
    list_per_page = 20

    fieldsets = (
        ('基本信息', {
            'fields': ('question_type', 'chapter', 'stem'),
        }),
        ('选项（根据题型填写对应选项）', {
            'fields': ('option_a', 'option_b', 'option_c', 'option_d', 'option_e', 'option_f'),
            'description': '单选题填写A~D；多选题填写A~F；判断题无需填写选项',
        }),
        ('答案与解析', {
            'fields': ('answer', 'analysis'),
        }),
    )

    # ── 自定义 URL ──
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-questions/',
                self.admin_site.admin_view(self.import_questions_view),
                name='questions_question_import',
            ),
        ]
        return custom_urls + urls

    # ── 批量导入题目视图 ──
    def import_questions_view(self, request):
        """
        批量导入题目：上传 xlsx 文件
        xlsx 列格式：
          题目ID | 题目类型 | 所属章 | 所属节 | 所属小节
          | 题干 | 选项A | 选项B | 选项C | 选项D | 选项E | 选项F | 选项G | 正确答案
        """
        context = {
            'title': '批量导入题目',
            'opts': self.model._meta,
            'has_change_permission': self.has_change_permission(request),
            'site_header': self.admin_site.site_header,
            'site_title': self.admin_site.site_title,
        }

        if openpyxl is None:
            messages.error(request, '缺少 openpyxl 库。请运行: pip install openpyxl')
            return render(request, 'admin/questions/question/import_questions.html', context)

        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, '请选择一个 xlsx 文件上传')
                return render(request, 'admin/questions/question/import_questions.html', context)

            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, '仅支持 .xlsx 格式文件')
                return render(request, 'admin/questions/question/import_questions.html', context)

            try:
                wb = openpyxl.load_workbook(excel_file, read_only=True)
                ws = wb.active
            except Exception as e:
                messages.error(request, f'文件读取失败: {e}')
                return render(request, 'admin/questions/question/import_questions.html', context)

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                messages.warning(request, '文件中没有数据行（仅有表头）')
                return render(request, 'admin/questions/question/import_questions.html', context)

            # ── 解析表头，定位各列位置 ──
            header = [str(c or '').strip() for c in rows[0]]
            expected_cols = ['题目类型', '题干', '正确答案']
            if not all(col in header for col in expected_cols):
                messages.error(
                    request,
                    f'表头格式不符合要求，必须包含列：题目类型、题干、正确答案。'
                    f'当前表头: {", ".join(header)}'
                )
                return render(request, 'admin/questions/question/import_questions.html', context)

            # 构建列索引映射
            all_cols = [
                '题目ID', '题目类型', '所属章', '所属节', '所属小节',
                '题干', '选项A', '选项B', '选项C', '选项D',
                '选项E', '选项F', '选项G', '正确答案',
            ]
            col_idx = {}
            for col in all_cols:
                if col in header:
                    col_idx[col] = header.index(col)

            # ── 预加载题型和章节缓存 ──
            question_type_cache = {}
            for qt in QuestionType.objects.all():
                question_type_cache[qt.name] = qt

            # 章节缓存：name -> Chapter 对象（导入过程中逐步填充）
            chapter_node_cache = {}

            def _get_or_create_chapter(zhang_name, jie_name='', xiaojie_name=''):
                """按章→节→小节层级查找或创建 Chapter 节点"""
                # 第1级：章
                zhang_key = (None, zhang_name)
                if zhang_key in chapter_node_cache:
                    zhang_node = chapter_node_cache[zhang_key]
                else:
                    zhang_node, _ = Chapter.objects.get_or_create(
                        parent=None, name=zhang_name,
                        defaults={'sort_order': 0, 'level_depth': 1},
                    )
                    chapter_node_cache[zhang_key] = zhang_node

                if not jie_name:
                    return zhang_node

                # 第2级：节
                jie_key = (zhang_node.pk, jie_name)
                if jie_key in chapter_node_cache:
                    jie_node = chapter_node_cache[jie_key]
                else:
                    jie_node, _ = Chapter.objects.get_or_create(
                        parent=zhang_node, name=jie_name,
                        defaults={'sort_order': 0, 'level_depth': 2},
                    )
                    chapter_node_cache[jie_key] = jie_node

                if not xiaojie_name:
                    return jie_node

                # 第3级：小节
                xiaojie_key = (jie_node.pk, xiaojie_name)
                if xiaojie_key in chapter_node_cache:
                    return chapter_node_cache[xiaojie_key]
                else:
                    xiaojie_node, _ = Chapter.objects.get_or_create(
                        parent=jie_node, name=xiaojie_name,
                        defaults={'sort_order': 0, 'level_depth': 3},
                    )
                    chapter_node_cache[xiaojie_key] = xiaojie_node
                    return xiaojie_node

            # ── 逐行导入 ──
            imported_count = 0
            skipped_rows = []
            total_rows = len(rows) - 1  # 减去表头

            for row_idx, row in enumerate(rows[1:], start=2):
                # 跳过全空行
                if all(c is None or str(c).strip() == '' for c in row):
                    continue

                row_num = row_idx

                try:
                    # ── 读取各字段 ──
                    # 题目ID（可选）
                    question_id_str = str(row[col_idx['题目ID']]).strip() if '题目ID' in col_idx else ''
                    question_id = int(question_id_str) if question_id_str and question_id_str.isdigit() else None

                    # 如果指定了题目ID且该题目已存在，跳过（不覆盖）
                    if question_id and Question.objects.filter(id=question_id).exists():
                        skipped_rows.append(f'第{row_num}行：题目ID={question_id} 已存在，跳过')
                        continue

                    # ── 题目类型 ──
                    type_name = str(row[col_idx['题目类型']] or '').strip()
                    if type_name not in question_type_cache:
                        skipped_rows.append(f'第{row_num}行：题目类型「{type_name}」不存在，请先添加该题型')
                        continue
                    question_type = question_type_cache[type_name]

                    # ── 所属章节 ──
                    zhang = str(row[col_idx.get('所属章', -1)] or '').strip() if '所属章' in col_idx else ''
                    jie = str(row[col_idx.get('所属节', -1)] or '').strip() if '所属节' in col_idx else ''
                    xiaojie = str(row[col_idx.get('所属小节', -1)] or '').strip() if '所属小节' in col_idx else ''

                    if not zhang:
                        skipped_rows.append(f'第{row_num}行：所属章为空，跳过')
                        continue

                    chapter = _get_or_create_chapter(zhang, jie, xiaojie)

                    # ── 题干 ──
                    stem = str(row[col_idx['题干']] or '').strip()
                    if not stem:
                        skipped_rows.append(f'第{row_num}行：题干为空，跳过')
                        continue

                    # ── 选项（A~G） ──
                    option_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
                    options = []
                    for letter in option_letters:
                        col_name = f'选项{letter}'
                        if col_name in col_idx:
                            value = str(row[col_idx[col_name]] or '').strip()
                            if value:
                                options.append(f'{letter}. {value}')

                    # 判断题特殊处理：如果选项列为空，使用默认选项
                    if question_type.code == 'judgment' and not options:
                        options = ['A. 正确', 'B. 错误']

                    # ── 正确答案 ──
                    raw_answer = str(row[col_idx['正确答案']] or '').strip()

                    # 根据题型转换答案格式
                    if question_type.code == 'judgment':
                        # 判断题：A→正确, B→错误
                        if raw_answer.upper() == 'A':
                            answer = '正确'
                        elif raw_answer.upper() == 'B':
                            answer = '错误'
                        else:
                            answer = raw_answer  # 直接使用文本
                    elif question_type.code == 'multi_choice':
                        # 多选题：将 "ABCDEFG" 转为 "A,B,C,D,E,F,G"
                        # 也可能是逗号分隔的格式
                        if ',' in raw_answer:
                            answer = raw_answer  # 已经是逗号分隔
                        else:
                            # 将连续字母转为逗号分隔
                            clean = raw_answer.upper().strip()
                            answer = ','.join(list(clean))
                    else:
                        # 单选题：直接使用字母
                        answer = raw_answer.upper().strip()

                    # ── 构建题目数据 ──
                    question_data = {
                        'question_type': question_type,
                        'chapter': chapter,
                        'stem': stem,
                        'options': options,
                        'answer': answer,
                    }

                    # 如果提供了题目ID，使用指定ID创建
                    if question_id:
                        question_data['id'] = question_id

                    Question.objects.create(**question_data)
                    imported_count += 1

                except Exception as e:
                    skipped_rows.append(f'第{row_num}行导入出错: {e}')
                    continue

            wb.close()

            # ── 结果反馈 ──
            if imported_count > 0:
                messages.success(request, f'✅ 成功导入 {imported_count}/{total_rows} 道题目！')
            if skipped_rows:
                messages.warning(request, f'⚠️ 以下 {len(skipped_rows)} 行被跳过：')
                for msg in skipped_rows[:15]:
                    messages.warning(request, msg)
                if len(skipped_rows) > 15:
                    messages.warning(request, f'... 还有 {len(skipped_rows) - 15} 行被跳过，请检查文件内容')

            return HttpResponseRedirect(reverse('admin:questions_question_changelist'))

        return render(request, 'admin/questions/question/import_questions.html', context)

    # ── 列表展示方法 ──
    def stem_short(self, obj):
        return obj.stem[:60] + '...' if len(obj.stem) > 60 else obj.stem
    stem_short.short_description = '题干'

    def answer_display(self, obj):
        return format_html('<strong>{}</strong>', obj.answer)
    answer_display.short_description = '答案'

    def get_fieldsets(self, request, obj=None):
        fieldsets = super().get_fieldsets(request, obj)
        if obj and obj.question_type and obj.question_type.code == 'judgment':
            fieldsets = (
                ('基本信息', {
                    'fields': ('question_type', 'chapter', 'stem'),
                }),
                ('判断题设置', {
                    'fields': ('answer',),
                    'description': '判断题答案只能填写"正确"或"错误"',
                }),
                ('解析', {
                    'fields': ('analysis',),
                }),
            )
        return fieldsets