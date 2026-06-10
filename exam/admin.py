# exam/admin.py

"""
考试核心 Admin 配置（T-03~T-12 完整版 + P0 状态字段修复）
- Exam：基本信息 Tab + 出题规则内联 + 考生名单 + 状态+及格线
- Candidate：CSV导入 + 搜索
- ExamPaper：只读详情页 + 作答明细 + 统计摘要
- T-11：成绩导出 Action
- T-12：Admin UI 只读控制
"""
import re
import csv
import io
import logging
from django.contrib import admin
from django import forms
from django.shortcuts import render, redirect
from django.urls import path, reverse
from django.http import HttpResponse
from django.contrib import messages
from django.template.response import TemplateResponse
from django.utils.html import format_html
from django.utils import timezone
from django.http import HttpResponseRedirect

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Exam, QuestionRule, Candidate, ExamCandidate,
    ExamPaper, ExamPaperQuestion, Answer
)
from .services.paper_generator import PaperGenerator, PaperGenerationError
from .services.grading import force_finish_exam_papers

logger = logging.getLogger('exam')



class QuestionRuleInline(admin.TabularInline):
    model = QuestionRule
    extra = 1
    min_num = 1
    verbose_name = '出题规则'
    verbose_name_plural = '出题规则'
    fields = ('chapter', 'question_type', 'question_count')
    autocomplete_fields = ['chapter', 'question_type']

    def has_change_permission(self, request, obj=None):
        if obj and obj.exam_papers.exists():
            return False
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.exam_papers.exists():
            return False
        return super().has_delete_permission(request, obj)


class ExamCandidateInline(admin.TabularInline):
    model = ExamCandidate
    extra = 0
    can_delete = False
    verbose_name = '考生'
    verbose_name_plural = '已导入考生'
    fields = ('candidate', 'exam_paper', 'created_at')
    readonly_fields = ('exam_paper', 'created_at')
    autocomplete_fields = ['candidate']

    def has_add_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        if obj and obj.exam_papers.exists():
            return False
        return super().has_change_permission(request, obj)


@admin.register(Exam)
class ExamAdmin(admin.ModelAdmin):
    change_list_template = 'admin/exam/exam/change_list.html'

    list_display = (
        'name', 'status', 'start_time', 'end_time', 'duration_minutes',
        'total_questions', 'total_score', 'passing_score', 'candidate_count',
        'paper_status', 'created_at',
    )
    list_filter = ('status', 'start_time')
    search_fields = ('name',)
    date_hierarchy = 'start_time'
    inlines = [QuestionRuleInline, ExamCandidateInline]

    fieldsets = (
        ('基本信息', {
            'fields': (
                'name', 'status',
                ('start_time', 'end_time'),
                ('duration_minutes',),
                ('total_questions', 'total_score', 'passing_score'),
            ),
        }),
    )

    actions = ['generate_papers', 'export_scores_xlsx', 'force_finish_expired']

    def get_readonly_fields(self, request, obj=None):
        readonly = super().get_readonly_fields(request, obj)
        if obj and obj.exam_papers.exists():
            return ['name', 'start_time', 'end_time', 'duration_minutes',
                    'total_questions', 'total_score', 'passing_score'] + list(readonly)
        return readonly

    def has_change_permission(self, request, obj=None):
        if obj and obj.exam_papers.exists():
            return True
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.exam_papers.exists():
            return False
        return super().has_delete_permission(request, obj)

    def candidate_count(self, obj):
        return obj.exam_candidates.count()
    candidate_count.short_description = '考生人数'

    def paper_status(self, obj):
        count = obj.exam_papers.count()
        if count == 0:
            return '❌ 未组卷'
        return f'✅ 已组卷（{count}份）'
    paper_status.short_description = '组卷状态'

    @admin.action(description='📄 生成试卷')
    def generate_papers(self, request, queryset):
        for exam in queryset:
            if exam.exam_papers.exists():
                self.message_user(request, f'考试「{exam.name}」已组卷', level='WARNING')
                continue
            try:
                generator = PaperGenerator(exam)
                generator.generate()
                self.message_user(
                    request,
                    f'✅ 考试「{exam.name}」组卷成功！共 {exam.exam_papers.count()} 份试卷',
                    level='SUCCESS',
                )
            except PaperGenerationError as e:
                self.message_user(
                    request,
                    f'❌ 考试「{exam.name}」组卷失败：{str(e)}',
                    level='ERROR',
                )

    @admin.action(description='📊 导出成绩XLSX')
    def export_scores_xlsx(self, request, queryset):
        """
        导出已结束考试的全部考生成绩为 .xlsx
        - 仅导出考试状态为 FINISHED 的考试
        - 每位考生一行：序号、姓名、身份证号、总分、及格、状态、交卷时间
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from django.utils import timezone
        import re

        # ── 1. 区分已结束 vs 未结束的考试 ──
        finished_exams = []  # 可以导出的
        skipped_names = []  # 要跳过的

        for exam in queryset:
            if exam.status == Exam.Status.FINISHED:
                finished_exams.append(exam)
            else:
                skipped_names.append(exam.name)

        if not finished_exams:
            self.message_user(
                request,
                '⚠️ 所选考试均未结束，无法导出成绩。'
                '请先结束考试后再导出。',
                level='WARNING',
            )
            return

        # ── 2. 创建工作簿 ──
        wb = openpyxl.Workbook()

        # 表头样式
        header_font = Font(bold=True, size=11)
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        headers = ['序号', '姓名', '身份证号', '总分', '及格', '状态', '交卷时间']

        total_rows = 0

        for exam in finished_exams:
            # 每个考试一个 Sheet，Sheet 名取考试名称（截断防超长）
            sheet_name = exam.name[:31]  # Excel Sheet 名最多 31 字符
            ws = wb.active if finished_exams.index(exam) == 0 else wb.create_sheet()
            ws.title = sheet_name

            # 写入表头
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border

            # ── 3. 查询该考试的全部考生（通过 ExamCandidate） ──
            candidates = (
                ExamCandidate.objects
                .filter(exam=exam)
                .select_related('candidate', 'exam_paper')
                .order_by('candidate__name')
            )

            row_num = 1
            for ec in candidates:
                row_num += 1

                # 确定状态、分数、交卷时间
                paper = ec.exam_paper
                if paper is None:
                    status_text = '缺考'
                    score = 0
                    submitted_at = ''
                elif paper.status == ExamPaper.Status.FINISHED:
                    status_text = '已交卷'
                    score = float(paper.total_score) if paper.total_score else 0
                    submitted_at = (
                        paper.submitted_at.strftime('%Y-%m-%d %H:%M:%S')
                        if paper.submitted_at else ''
                    )
                else:
                    status_text = '缺考'
                    score = 0
                    submitted_at = ''

                passed = '是' if score >= float(exam.passing_score) else '否'

                values = [
                    row_num - 1,
                    ec.candidate.name,
                    ec.candidate.id_card,
                    score,
                    passed,
                    status_text,
                    submitted_at,
                ]

                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.border = thin_border

                total_rows += 1

            # 自适应列宽（取表头和数据的最大长度）
            for col_idx in range(1, len(headers) + 1):
                max_len = len(str(ws.cell(row=1, column=col_idx).value))
                for r in range(2, row_num + 1):
                    cell_val = ws.cell(row=r, column=col_idx).value
                    if cell_val is not None:
                        # 中文字符按 2 倍宽度估算
                        val_len = sum(2 if ord(c) > 127 else 1 for c in str(cell_val))
                        max_len = max(max_len, val_len)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

        # ── 4. 生成响应 ──
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        # 文件名
        timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        if len(finished_exams) == 1:
            # 单场考试：用考试名称
            safe_name = re.sub(r'[\\/:*?"<>|]', '_', finished_exams[0].name)
            filename = f'{safe_name}_{timestamp}.xlsx'
        else:
            filename = f'考试成绩汇总_{timestamp}.xlsx'

        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)

        # ── 5. 消息提示 ──
        msg_parts = [f'✅ 已导出 {total_rows} 条成绩记录']
        if skipped_names:
            msg_parts.append(f'⚠️ 以下考试尚未结束，已跳过：{", ".join(skipped_names)}')
        self.message_user(request, ' | '.join(msg_parts), level='SUCCESS')

        return response

    export_scores_xlsx.short_description = '📊 导出成绩XLSX'

    @admin.action(description='⏰ 结束考试')
    def force_finish_expired(self, request, queryset):
        count = 0
        exam_names = []
        for exam in queryset:
            c = force_finish_exam_papers(exam)
            if c > 0:
                count += c
                exam_names.append(exam.name)
        if count > 0:
            self.message_user(
                request,
                f'⏰ 已强制交卷并批改 {count} 份试卷'
                f'（考试：{", ".join(exam_names)}）',
            )
        else:
            self.message_user(request, '所选考试没有进行中的试卷需要处理')
    force_finish_expired.short_description = '⏰ 结束考试'

    # ─── save_model 只做保存，不做校验 ───
    def save_model(self, request, obj, form, change):
        """仅保存 Exam 主干，校验逻辑移至 save_related（确保 inline 已保存）"""
        super().save_model(request, obj, form, change)

    # ─── save_related 做校验（此时 inline 已保存） ───
    def save_related(self, request, form, formsets, change):
        """保存 inline 后执行校验"""
        from questions.models import Question

        super().save_related(request, form, formsets, change)

        obj = form.instance
        rules = list(obj.question_rules.all().select_related('chapter', 'question_type'))

        if len(rules) == 0:
            messages.error(request, '请先添加至少一条出题规则')
            return

        total_from_rules = sum(r.question_count for r in rules)
        total_score_from_rules = sum(
            r.question_count * r.question_type.score for r in rules
        )

        has_error = False

        # ── 校验总题数 ──
        if total_from_rules != obj.total_questions:
            has_error = True
            messages.error(
                request,
                f'❌ 总题数不一致：考试设置 {obj.total_questions} 题，'
                f'但出题规则合计 {total_from_rules} 题'
                f'（差额 {total_from_rules - obj.total_questions} 题）'
            )

        # ── 校验总分 ──
        if total_score_from_rules != obj.total_score:
            has_error = True
            messages.error(
                request,
                f'❌ 总分不一致：考试设置 {obj.total_score} 分，'
                f'但出题规则合计 {total_score_from_rules} 分'
                f'（差额 {total_score_from_rules - obj.total_score} 分）'
            )

        # ── 逐条校验：递归查找章节及所有子章节的题目 ──
        shortage_rules = []
        for r in rules:
            actual_count = self._count_questions_in_chapter_tree(
                r.chapter, r.question_type
            )
            if actual_count < r.question_count:
                shortage_rules.append(
                    f'章节「{r.chapter.name}」+ {r.question_type.name}：'
                    f'需要 {r.question_count} 题，题库（含子章节）仅有 {actual_count} 题，'
                    f'缺少 {r.question_count - actual_count} 题'
                )

        if shortage_rules:
            has_error = True
            for msg in shortage_rules:
                messages.warning(request, f'⚠️ {msg}')

        if has_error:
            return

        messages.success(request, '✅ 考试保存成功，待考生名单确认后触发组卷')

    # ── 递归获取章节及其所有子章节的题目数量 ──
    def _count_questions_in_chapter_tree(self, chapter, question_type):
        """统计指定章节及其所有子章节中某题型的题目数量"""
        from questions.models import Question, Chapter as ChapterModel

        def _get_descendant_ids(ch):
            ids = {ch.id}
            children = ChapterModel.objects.filter(parent=ch).only('id')
            for child in children:
                ids.update(_get_descendant_ids(child))
            return ids

        chapter_ids = _get_descendant_ids(chapter)
        return Question.objects.filter(
            chapter_id__in=chapter_ids,
            question_type=question_type,
        ).count()

    # ── 批量导入（考试+规则+考生） ──
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-exams/',
                self.admin_site.admin_view(self.import_exams_view),
                name='exam_exam_import',
            ),
        ]
        return custom_urls + urls

    def import_exams_view(self, request):
        """批量导入考试（含出题规则 + 考生名单）"""
        context = {
            'title': '批量导入考试',
            'opts': self.model._meta,
            'has_change_permission': self.has_change_permission(request),
        }

        if openpyxl is None:
            messages.error(request, '缺少 openpyxl 库。请运行: pip install openpyxl')
            return render(request, 'admin/exam/exam/import_exams.html', context)

        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, '请选择一个 xlsx 文件上传')
                return render(request, 'admin/exam/exam/import_exams.html', context)

            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, '仅支持 .xlsx 格式文件')
                return render(request, 'admin/exam/exam/import_exams.html', context)

            try:
                wb = openpyxl.load_workbook(excel_file)
            except Exception as e:
                messages.error(request, f'文件读取失败: {e}')
                return render(request, 'admin/exam/exam/import_exams.html', context)

            from .services.exam_import import process_exam_import
            result = process_exam_import(wb, request)

            if not result.get('success'):
                return render(request, 'admin/exam/exam/import_exams.html', context)

            parts = []
            if result['exam_created']:
                parts.append(f'考试: ✅ {result["exam_created"]} 个')
            if result['exam_skipped']:
                parts.append(f'考试: ⏭️ {result["exam_skipped"]} 个已存在跳过')
            if result['rule_created']:
                parts.append(f'出题规则: ✅ {result["rule_created"]} 条')
            if result['rule_skipped']:
                parts.append(f'出题规则: ⏭️ {result["rule_skipped"]} 条跳过')
            if result['candidate_added']:
                parts.append(f'考生: ✅ {result["candidate_added"]} 人')
            if result['candidate_skipped']:
                parts.append(f'考生: ⏭️ {result["candidate_skipped"]} 人跳过')

            if parts:
                messages.success(request, ' | '.join(parts))
            else:
                messages.warning(request, '没有新的数据被导入，请检查文件内容')

            return HttpResponseRedirect(reverse('admin:exam_exam_changelist'))

        return render(request, 'admin/exam/exam/import_exams.html', context)


@admin.register(QuestionRule)
class QuestionRuleAdmin(admin.ModelAdmin):
    list_display = ('exam', 'chapter', 'question_type', 'question_count', 'rule_score')
    list_filter = ('exam',)
    search_fields = ('exam__name',)

    def rule_score(self, obj):
        return obj.question_count * obj.question_type.score
    rule_score.short_description = '小计分数'

    def has_change_permission(self, request, obj=None):
        if obj and obj.exam.exam_papers.exists():
            return False
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.exam.exam_papers.exists():
            return False
        return super().has_delete_permission(request, obj)


class CsvImportForm(forms.Form):
    csv_file = forms.FileField(
        label='选择 CSV 文件',
        help_text='CSV 格式：第一列身份证号，第二列姓名（无表头行）',
    )


@admin.register(Candidate)
class CandidateAdmin(admin.ModelAdmin):
    list_display = ('name', 'id_card', 'exam_count', 'created_at')
    search_fields = ('name', 'id_card')

    def exam_count(self, obj):
        return obj.exam_candidates.count()
    exam_count.short_description = '参与考试数'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-csv/',
                self.admin_site.admin_view(self.csv_import_view),
                name='exam_candidate_csv_import',
            ),
        ]
        return custom_urls + urls

    def csv_import_view(self, request):
        if request.method == 'POST':
            form = CsvImportForm(request.POST, request.FILES)
            if 'confirm' in request.POST:
                return self._do_import(request, form)
            if form.is_valid():
                csv_file = request.FILES['csv_file']
                decoded_file = csv_file.read().decode('utf-8-sig')
                reader = csv.reader(io.StringIO(decoded_file))
                rows = []
                errors = []
                for line_num, row in enumerate(reader, start=1):
                    if not row or not any(row):
                        continue
                    if len(row) < 2:
                        errors.append(f'第 {line_num} 行：格式错误，需要至少两列')
                        continue
                    id_card = row[0].strip().upper()
                    name = row[1].strip()
                    if not id_card or not name:
                        errors.append(f'第 {line_num} 行：身份证号或姓名为空')
                        continue
                    import re
                    if not re.match(
                        r'^[1-9]\d{5}(19|20)\d{2}(0[1-9]|1[0-2])'
                        r'(0[1-9]|[12]\d|3[01])\d{3}[\dX]$', id_card
                    ):
                        errors.append(f'第 {line_num} 行：身份证号格式不正确（{id_card}）')
                        continue
                    rows.append({'id_card': id_card, 'name': name})
                if not rows:
                    messages.error(request, 'CSV 文件为空或格式无效')
                    return render(
                        request, 'admin/exam/candidate/csv_import.html',
                        {'form': form, 'title': '导入考生'},
                    )
                preview = rows[:10]
                request.session['csv_import_rows'] = rows
                return render(request, 'admin/exam/candidate/csv_import.html', {
                    'form': form, 'title': '导入考生 - 确认', 'preview': preview,
                    'total': len(rows), 'error_count': len(errors),
                    'error_details': errors[:10], 'show_confirm': True,
                })
        else:
            form = CsvImportForm()
        return render(
            request, 'admin/exam/candidate/csv_import.html',
            {'form': form, 'title': '导入考生'},
        )

    def _do_import(self, request, form):
        rows = request.session.pop('csv_import_rows', [])
        if not rows:
            messages.error(request, '导入数据已丢失，请重新上传')
            return redirect('..')
        success_count = 0
        fail_count = 0
        fail_details = []
        for row in rows:
            try:
                Candidate.objects.get_or_create(
                    id_card=row['id_card'],
                    defaults={'name': row['name']},
                )
                success_count += 1
            except Exception as e:
                fail_count += 1
                fail_details.append(f'{row["id_card"]}（{row["name"]}）：{str(e)}')
        logger.info(
            'CSV导入完成 | 总行=%d 成功=%d 失败=%d',
            len(rows), success_count, fail_count,
        )
        msg = f'✅ 成功导入 {success_count} 人'
        if fail_count > 0:
            msg += f'，{fail_count} 人导入失败'
            for detail in fail_details[:5]:
                msg += f'\n  - {detail}'
            messages.warning(request, msg)
        else:
            messages.success(request, msg)
        return redirect('..')


class AnswerInline(admin.StackedInline):
    model = Answer
    extra = 0
    can_delete = False
    fields = ('selected_answer', 'is_correct', 'score', 'answered_at')
    readonly_fields = ('selected_answer', 'is_correct', 'score', 'answered_at')

    def has_add_permission(self, request, obj=None):
        return False


class ExamPaperQuestionInline(admin.TabularInline):
    model = ExamPaperQuestion
    extra = 0
    can_delete = False
    fields = ('sort_order', 'question', 'shuffled_options', 'original_answer', 'score')
    readonly_fields = ('sort_order', 'question', 'shuffled_options', 'original_answer', 'score')
    show_change_link = True

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(ExamPaper)
class ExamPaperAdmin(admin.ModelAdmin):
    list_display = ('id', 'exam', 'candidate', 'status', 'total_score', 'started_at', 'submitted_at')
    list_filter = ('status', 'exam')
    search_fields = ('candidate__name', 'candidate__id_card')
    inlines = [ExamPaperQuestionInline]
    readonly_fields = (
        'exam', 'candidate', 'status', 'total_score',
        'started_at', 'submitted_at', 'answer_summary',
    )
    fieldsets = (
        ('试卷信息', {
            'fields': ('exam', 'candidate', 'status', ('total_score',), ('started_at', 'submitted_at')),
        }),
        ('统计摘要', {
            'fields': ('answer_summary',),
            'description': '作答统计（自动计算）',
        }),
    )

    def answer_summary(self, obj):
        answers = Answer.objects.filter(exam_paper_question__exam_paper=obj)
        total = answers.count()
        answered = answers.exclude(selected_answer=None).count()
        correct = answers.filter(is_correct=True).count()
        wrong = answers.filter(is_correct=False).count()
        unanswered = answers.filter(selected_answer=None).count()
        return (
            f'总题数：{total} ｜ 已作答：{answered} ｜ '
            f'正确：{correct} ｜ 错误：{wrong} ｜ 未作答：{unanswered}'
        )
    answer_summary.short_description = '作答统计'

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(ExamPaperQuestion)
class ExamPaperQuestionAdmin(admin.ModelAdmin):
    list_display = ('id', 'exam_paper', 'sort_order', 'question', 'score')
    search_fields = ('exam_paper__candidate__name',)
    inlines = [AnswerInline]

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

# # exam/admin.py
#
# """
# 考试核心 Admin 配置（T-03 + T-04）
# - Exam：基本信息 Tab + 出题规则内联 + 考生名单内联 + 批量导入
# - Candidate：搜索
# - ExamPaper：详情页展示完整作答明细
# - 已组卷考试只读保护（P1-004）
# """
# from django.contrib import admin
# from django import forms
# from django.shortcuts import redirect
# from django.urls import reverse, path
# from django.shortcuts import render
# from django.contrib import messages
# from django.http import HttpResponseRedirect
#
# from .models import (
#     Exam, QuestionRule, Candidate, ExamCandidate,
#     ExamPaper, ExamPaperQuestion, Answer
# )
#
# from .services.exam_import import process_exam_import
#
# # ─── 批量导入依赖 ─────────────────────────────
# try:
#     import openpyxl
#     from openpyxl.utils import get_column_letter
# except ImportError:
#     openpyxl = None
#
#
# class QuestionRuleInline(admin.TabularInline):
#     model = QuestionRule
#     extra = 1
#     min_num = 1
#     verbose_name = '出题规则'
#     verbose_name_plural = '出题规则'
#     fields = ('chapter', 'question_type', 'question_count')
#     autocomplete_fields = ['chapter', 'question_type']
#
#
# class ExamCandidateInline(admin.TabularInline):
#     model = ExamCandidate
#     extra = 0
#     can_delete = False
#     verbose_name = '考生'
#     verbose_name_plural = '已导入考生'
#     fields = ('candidate', 'exam_paper', 'created_at')
#     readonly_fields = ('exam_paper', 'created_at')
#     autocomplete_fields = ['candidate']
#
#     def has_add_permission(self, request, obj=None):
#         return False
#
#
# # ═══════════════════════════════════════════════
# # Exam Admin — 考试管理 + 批量导入
# # ═══════════════════════════════════════════════
#
# @admin.register(Exam)
# class ExamAdmin(admin.ModelAdmin):
#     change_list_template = 'admin/exam/exam/change_list.html'
#
#     list_display = ('name', 'start_time', 'end_time', 'duration_minutes',
#                     'total_questions', 'total_score', 'candidate_count',
#                     'paper_status', 'created_at')
#     list_filter = ('start_time',)
#     search_fields = ('name',)
#     date_hierarchy = 'start_time'
#     inlines = [QuestionRuleInline, ExamCandidateInline]
#
#     fieldsets = (
#         ('基本信息', {
#             'fields': ('name', ('start_time', 'end_time'), ('duration_minutes',),
#                        ('total_questions', 'total_score')),
#         }),
#     )
#
#     def get_urls(self):
#         urls = super().get_urls()
#         custom_urls = [
#             path(
#                 'import-exams/',
#                 self.admin_site.admin_view(self.import_exams_view),
#                 name='exam_exam_import',
#             ),
#         ]
#         return custom_urls + urls
#
#     def import_exams_view(self, request):
#         """批量导入考试：上传 xlsx → 交 process_exam_import 处理"""
#         context = {
#             'title': '批量导入考试',
#             'opts': self.model._meta,
#             'has_change_permission': self.has_change_permission(request),
#             'site_header': self.admin_site.site_header,
#             'site_title': self.admin_site.site_title,
#         }
#
#         if openpyxl is None:
#             messages.error(request, '缺少 openpyxl 库。请运行: pip install openpyxl')
#             return render(request, 'admin/exam/exam/import_exams.html', context)
#
#         if request.method == 'POST':
#             excel_file = request.FILES.get('excel_file')
#             if not excel_file:
#                 messages.error(request, '请选择一个 xlsx 文件上传')
#                 return render(request, 'admin/exam/exam/import_exams.html', context)
#
#             if not excel_file.name.endswith(('.xlsx', '.xls')):
#                 messages.error(request, '仅支持 .xlsx 格式文件')
#                 return render(request, 'admin/exam/exam/import_exams.html', context)
#
#             try:
#                 wb = openpyxl.load_workbook(excel_file)
#             except Exception as e:
#                 messages.error(request, f'文件读取失败: {e}')
#                 return render(request, 'admin/exam/exam/import_exams.html', context)
#
#             result = process_exam_import(wb, request)
#
#             if not result.get('success'):
#                 return render(request, 'admin/exam/exam/import_exams.html', context)
#
#             parts = []
#             if result['exam_created']:
#                 parts.append(f'考试: ✅ {result["exam_created"]} 个')
#             if result['exam_skipped']:
#                 parts.append(f'考试: ⏭️ {result["exam_skipped"]} 个已存在跳过')
#             if result['rule_created']:
#                 parts.append(f'出题规则: ✅ {result["rule_created"]} 条')
#             if result['rule_skipped']:
#                 parts.append(f'出题规则: ⏭️ {result["rule_skipped"]} 条跳过')
#             if result['candidate_added']:
#                 parts.append(f'考生: ✅ {result["candidate_added"]} 人')
#             if result['candidate_skipped']:
#                 parts.append(f'考生: ⏭️ {result["candidate_skipped"]} 人跳过')
#
#             if parts:
#                 messages.success(request, ' | '.join(parts))
#             else:
#                 messages.warning(request, '没有新的数据被导入，请检查文件内容')
#
#             return HttpResponseRedirect(reverse('admin:exam_exam_changelist'))
#
#         return render(request, 'admin/exam/exam/import_exams.html', context)
#
#     # ── 统计考生人数 ──
#     def candidate_count(self, obj):
#         return obj.exam_candidates.count()
#
#     candidate_count.short_description = '考生人数'
#
#     # ── 组卷状态显示 ──
#     def paper_status(self, obj):
#         count = obj.exam_papers.count()
#         if count == 0:
#             return '❌ 未组卷'
#         return f'✅ 已组卷（{count}份）'
#
#     paper_status.short_description = '组卷状态'
#
#     # ── 注册 Admin Action ──
#     actions = ['generate_papers']
#
#     @admin.action(description='📄 生成试卷（为当前考试所有考生生成试卷）')
#     def generate_papers(self, request, queryset):
#         from exam.services.paper_generator import PaperGenerator, PaperGenerationError
#
#         for exam in queryset:
#             if exam.exam_papers.exists():
#                 self.message_user(
#                     request,
#                     f'考试「{exam.name}」已组卷，如需重新组卷请先清空现有试卷',
#                     level='WARNING'
#                 )
#                 continue
#
#             try:
#                 generator = PaperGenerator(exam)
#                 generator.generate()
#                 self.message_user(
#                     request,
#                     f'✅ 考试「{exam.name}」组卷成功！'
#                     f'共 {exam.exam_papers.count()} 份试卷',
#                     level='SUCCESS'
#                 )
#             except PaperGenerationError as e:
#                 self.message_user(
#                     request,
#                     f'❌ 考试「{exam.name}」组卷失败：{str(e)}',
#                     level='ERROR'
#                 )
#
#     # ─── save_model 只做保存，不做校验 ───
#     def save_model(self, request, obj, form, change):
#         """仅保存 Exam 主干，校验逻辑移至 save_related（确保 inline 已保存）"""
#         super().save_model(request, obj, form, change)
#
#     # ─── save_related 做校验（此时 inline 已保存） ───
#     def save_related(self, request, form, formsets, change):
#         """保存 inline 后执行校验"""
#         from questions.models import Question
#
#         super().save_related(request, form, formsets, change)
#
#         obj = form.instance
#         rules = list(obj.question_rules.all().select_related('chapter', 'question_type'))
#
#         if len(rules) == 0:
#             messages.error(request, '请先添加至少一条出题规则')
#             return
#
#         total_from_rules = sum(r.question_count for r in rules)
#         total_score_from_rules = sum(
#             r.question_count * r.question_type.score for r in rules
#         )
#
#         has_error = False
#
#         # ── 校验总题数 ──
#         if total_from_rules != obj.total_questions:
#             has_error = True
#             messages.error(
#                 request,
#                 f'❌ 总题数不一致：考试设置 {obj.total_questions} 题，'
#                 f'但出题规则合计 {total_from_rules} 题'
#                 f'（差额 {total_from_rules - obj.total_questions} 题）'
#             )
#
#         # ── 校验总分 ──
#         if total_score_from_rules != obj.total_score:
#             has_error = True
#             messages.error(
#                 request,
#                 f'❌ 总分不一致：考试设置 {obj.total_score} 分，'
#                 f'但出题规则合计 {total_score_from_rules} 分'
#                 f'（差额 {total_score_from_rules - obj.total_score} 分）'
#             )
#
#         # ── 逐条校验：递归查找章节及所有子章节的题目 ──
#         shortage_rules = []
#         for r in rules:
#             actual_count = self._count_questions_in_chapter_tree(
#                 r.chapter, r.question_type
#             )
#             if actual_count < r.question_count:
#                 shortage_rules.append(
#                     f'章节「{r.chapter.name}」+ {r.question_type.name}：'
#                     f'需要 {r.question_count} 题，题库（含子章节）仅有 {actual_count} 题，'
#                     f'缺少 {r.question_count - actual_count} 题'
#                 )
#
#         if shortage_rules:
#             has_error = True
#             for msg in shortage_rules:
#                 messages.warning(request, f'⚠️ {msg}')
#
#         if has_error:
#             return
#
#         messages.success(request, '✅ 考试保存成功，待考生名单确认后触发组卷')
#
#     # ── 递归获取章节及其所有子章节的题目数量 ──
#     def _count_questions_in_chapter_tree(self, chapter, question_type):
#         """统计指定章节及其所有子章节中某题型的题目数量"""
#         from questions.models import Question, Chapter as ChapterModel
#
#         def _get_descendant_ids(ch):
#             ids = {ch.id}
#             children = ChapterModel.objects.filter(parent=ch).only('id')
#             for child in children:
#                 ids.update(_get_descendant_ids(child))
#             return ids
#
#         chapter_ids = _get_descendant_ids(chapter)
#         return Question.objects.filter(
#             chapter_id__in=chapter_ids,
#             question_type=question_type,
#         ).count()
#
#
# @admin.register(QuestionRule)
# class QuestionRuleAdmin(admin.ModelAdmin):
#     list_display = ('exam', 'chapter', 'question_type', 'question_count', 'rule_score')
#     list_filter = ('exam',)
#     search_fields = ('exam__name',)
#
#     def rule_score(self, obj):
#         return obj.question_count * obj.question_type.score
#     rule_score.short_description = '小计分数'
#
#
# @admin.register(Candidate)
# class CandidateAdmin(admin.ModelAdmin):
#     list_display = ('name', 'id_card', 'created_at')
#     search_fields = ('name', 'id_card')
#     actions = []
#
#
# class AnswerInline(admin.StackedInline):
#     model = Answer
#     extra = 0
#     can_delete = False
#     fields = ('selected_answer', 'is_correct', 'score', 'answered_at')
#     readonly_fields = ('selected_answer', 'is_correct', 'score', 'answered_at')
#
#     def has_add_permission(self, request, obj=None):
#         return False
#
#
# class ExamPaperQuestionInline(admin.TabularInline):
#     model = ExamPaperQuestion
#     extra = 0
#     can_delete = False
#     fields = ('sort_order', 'question', 'shuffled_options', 'original_answer', 'score')
#     readonly_fields = ('sort_order', 'question', 'shuffled_options', 'original_answer', 'score')
#     show_change_link = True
#
#     def has_add_permission(self, request, obj=None):
#         return False
#
#     def has_delete_permission(self, request, obj=None):
#         return False
#
# @admin.register(ExamPaper)
# class ExamPaperAdmin(admin.ModelAdmin):
#     list_display = ('id', 'exam', 'candidate', 'status', 'total_score',
#                     'started_at', 'submitted_at')
#     list_filter = ('status', 'exam')
#     search_fields = ('candidate__name', 'candidate__id_card')
#     inlines = [ExamPaperQuestionInline]
#
#     fieldsets = (
#         ('试卷信息', {
#             'fields': ('exam', 'candidate', 'status',
#                        ('total_score',), ('started_at', 'submitted_at')),
#         }),
#         ('统计摘要', {
#             'fields': ('answer_summary',),
#             'description': '作答统计（自动计算）',
#         }),
#     )
#
#     # ── 永远只读的字段（方法/关系，不可编辑） ──
#     readonly_fields = ('exam', 'candidate', 'answer_summary')
#
#     def get_readonly_fields(self, request, obj=None):
#         """
#         权限控制：
#         - 超级管理员 / 有 change_exampaper 权限 → status/score/时间可编辑
#         - 其他用户 → 全只读
#         """
#         base_readonly = self.readonly_fields
#         if request.user.is_superuser or request.user.has_perm('exam.change_exampaper'):
#             return base_readonly
#         return base_readonly + ('status', 'total_score', 'started_at', 'submitted_at')
#
#     def has_add_permission(self, request, obj=None):
#         return request.user.is_superuser
#
#     def has_delete_permission(self, request, obj=None):
#         return request.user.is_superuser
#
#     def answer_summary(self, obj):
#         answers = Answer.objects.filter(exam_paper_question__exam_paper=obj)
#         total = answers.count()
#         answered = answers.exclude(selected_answer=None).count()
#         correct = answers.filter(is_correct=True).count()
#         wrong = answers.filter(is_correct=False).count()
#         unanswered = answers.filter(selected_answer=None).count()
#         return f'总题数：{total} ｜ 已作答：{answered} ｜ 正确：{correct} ｜ 错误：{wrong} ｜ 未作答：{unanswered}'
#     answer_summary.short_description = '作答统计'
#
#     def has_add_permission(self, request, obj=None):
#         return request.user.is_superuser
#
#     def has_delete_permission(self, request, obj=None):
#         return request.user.is_superuser